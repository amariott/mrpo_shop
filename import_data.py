import hashlib
import shutil
import sqlite3
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).parent
DB_PATH = ROOT / "database.db"

PRODUCTS_XLSX = ROOT / "Tovar.xlsx"
USERS_XLSX = ROOT / "user_import.xlsx"
ORDERS_XLSX = ROOT / "Заказ_import.xlsx"
PICKUP_POINTS_XLSX = ROOT / "Пункты выдачи_import.xlsx"

PRODUCT_IMAGES_DIR = ROOT / "static" / "images" / "products"

ALLOWED_TABLES = {
    "roles",
    "categories",
    "manufacturers",
    "suppliers",
    "order_statuses",
}


def get_connection():
    connection = sqlite3.connect(DB_PATH)
    connection.row_factory = sqlite3.Row
    connection.execute("PRAGMA foreign_keys = ON")
    return connection


def normalize_text(value):
    if value is None:
        return ""
    return str(value).strip()


def normalize_date(value):
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return normalize_text(value)


def hash_password(password):
    return hashlib.sha256(password.encode("utf-8")).hexdigest()


def row_is_empty(values):
    return not any(normalize_text(value) for value in values)


def get_or_create_id(connection, table_name, value):
    if table_name not in ALLOWED_TABLES:
        raise ValueError("Недопустимое имя таблицы")

    cursor = connection.execute(
        f"SELECT id FROM {table_name} WHERE name = ?",
        (value,),
    )
    row = cursor.fetchone()

    if row is not None:
        return row["id"]

    cursor = connection.execute(
        f"INSERT INTO {table_name} (name) VALUES (?)",
        (value,),
    )
    return cursor.lastrowid


def get_user_id_by_full_name(connection, full_name):
    rows = connection.execute(
        "SELECT id FROM users WHERE full_name = ?",
        (full_name,),
    ).fetchall()

    if not rows:
        raise ValueError(f"Пользователь не найден: {full_name}")

    return rows[0]["id"]


def get_product_id_by_article(connection, article):
    row = connection.execute(
        "SELECT id FROM products WHERE article = ?",
        (article,),
    ).fetchone()

    if row is None:
        raise ValueError(f"Товар не найден: {article}")

    return row["id"]


def copy_product_photo(file_name):
    file_name = normalize_text(file_name)

    if not file_name:
        return None

    source_path = ROOT / file_name

    if not source_path.exists():
        return None

    PRODUCT_IMAGES_DIR.mkdir(parents=True, exist_ok=True)
    target_path = PRODUCT_IMAGES_DIR / file_name
    shutil.copy2(source_path, target_path)

    return f"images/products/{file_name}"


def parse_order_items(raw_value):
    parts = [
        part.strip()
        for part in normalize_text(raw_value).split(",")
        if part.strip()
    ]

    if len(parts) % 2 != 0:
        raise ValueError(f"Неверный состав заказа: {raw_value}")

    items = []

    for index in range(0, len(parts), 2):
        article = parts[index]
        quantity = int(parts[index + 1])
        items.append((article, quantity))

    return items


def import_users(connection):
    worksheet = load_workbook(USERS_XLSX, data_only=True).active

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row_is_empty(row):
            continue

        role_name, full_name, login, password = row
        role_id = get_or_create_id(
            connection,
            "roles",
            normalize_text(role_name),
        )

        connection.execute(
            """
            INSERT INTO users (role_id, full_name, login, password_hash)
            VALUES (?, ?, ?, ?)
            """,
            (
                role_id,
                normalize_text(full_name),
                normalize_text(login),
                hash_password(normalize_text(password)),
            ),
        )


def import_products(connection):
    worksheet = load_workbook(PRODUCTS_XLSX, data_only=True).active

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row_is_empty(row):
            continue

        (
            article,
            name,
            unit,
            price,
            supplier_name,
            manufacturer_name,
            category_name,
            discount,
            stock,
            description,
            photo_name,
        ) = row

        supplier_id = get_or_create_id(
            connection,
            "suppliers",
            normalize_text(supplier_name),
        )
        manufacturer_id = get_or_create_id(
            connection,
            "manufacturers",
            normalize_text(manufacturer_name),
        )
        category_id = get_or_create_id(
            connection,
            "categories",
            normalize_text(category_name),
        )
        photo_path = copy_product_photo(photo_name)

        connection.execute(
            """
            INSERT INTO products (
                article,
                name,
                unit,
                price,
                supplier_id,
                manufacturer_id,
                category_id,
                discount,
                stock,
                description,
                photo_path
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                normalize_text(article),
                normalize_text(name),
                normalize_text(unit),
                float(price or 0),
                supplier_id,
                manufacturer_id,
                category_id,
                int(discount or 0),
                int(stock or 0),
                normalize_text(description),
                photo_path,
            ),
        )


def import_pickup_points(connection):
    worksheet = load_workbook(PICKUP_POINTS_XLSX, data_only=True).active
    addresses = []

    first_cell = normalize_text(worksheet.cell(1, 1).value)
    if first_cell:
        addresses.append(first_cell)

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row_is_empty(row):
            continue

        address = normalize_text(row[0])
        if address:
            addresses.append(address)

    for point_id, address in enumerate(addresses, start=1):
        connection.execute(
            "INSERT INTO pickup_points (id, address) VALUES (?, ?)",
            (point_id, address),
        )


def import_orders(connection):
    worksheet = load_workbook(ORDERS_XLSX, data_only=True).active

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row_is_empty(row[:8]):
            continue

        (
            order_id,
            raw_items,
            order_date,
            delivery_date,
            pickup_point_id,
            client_full_name,
            pickup_code,
            status_name,
            *_
        ) = row

        status_id = get_or_create_id(
            connection,
            "order_statuses",
            normalize_text(status_name),
        )
        user_id = get_user_id_by_full_name(
            connection,
            normalize_text(client_full_name),
        )

        connection.execute(
            """
            INSERT INTO orders (
                id,
                user_id,
                pickup_point_id,
                pickup_code,
                status_id,
                order_date,
                delivery_date
            )
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (
                int(order_id),
                user_id,
                int(pickup_point_id),
                normalize_text(pickup_code),
                status_id,
                normalize_date(order_date),
                normalize_date(delivery_date),
            ),
        )

        for article, quantity in parse_order_items(raw_items):
            product_id = get_product_id_by_article(connection, article)

            connection.execute(
                """
                INSERT INTO order_items (order_id, product_id, quantity)
                VALUES (?, ?, ?)
                """,
                (
                    int(order_id),
                    product_id,
                    quantity,
                ),
            )


def main():
    connection = get_connection()

    try:
        import_users(connection)
        import_products(connection)
        import_pickup_points(connection)
        import_orders(connection)
        connection.commit()
        print("Импорт завершен.")
    except Exception:
        connection.rollback()
        raise
    finally:
        connection.close()


if __name__ == "__main__":
    main()
